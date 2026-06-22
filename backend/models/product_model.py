"""
Product Model — data access layer.

Priority:
  1. Query Google Sheet API (GOOGLE_SHEET_URL in .env)
  2. Graceful degradation → mock-product.xlsx
"""
import os
import re
import requests
from typing import Optional
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

GOOGLE_SHEET_URL = os.getenv("GOOGLE_SHEET_URL", "")
MOCK_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "mock-product.xlsx")

# Cache mock data in memory after first load
_mock_cache: dict[str, dict] = {}


def _load_mock() -> dict[str, dict]:
    global _mock_cache
    if _mock_cache:
        return _mock_cache
    wb = load_workbook(MOCK_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        sku = record.get("SKU")
        if sku:
            _mock_cache[sku] = record
    wb.close()
    return _mock_cache


def _parse_float(value: Optional[str]) -> float:
    """Strip units like 'kg', 'g', 'mm', 'mm³' and return float."""
    if value is None:
        return 0.0
    cleaned = re.sub(r"[^\d.]", "", str(value))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _normalise(raw: dict) -> dict:
    """Return a clean product dict with numeric fields."""
    return {
        "sku": raw.get("SKU", ""),
        "product_name": raw.get("ProductName", ""),
        "rrp": _parse_float(raw.get("RRP")),
        "weight_g": _parse_float(raw.get("weight")),
        "volumetric_gross_weight_kg": _parse_float(raw.get("Volumetric_GrossWeight")),
        "length_mm": _parse_float(raw.get("length")),
        "width_mm": _parse_float(raw.get("width")),
        "height_mm": _parse_float(raw.get("height")),
        "volume_mm3": _parse_float(raw.get("volume")),
    }


def get_product_by_sku(sku: str) -> Optional[dict]:
    """Fetch product from Google Sheet API, fall back to mock on failure."""
    if GOOGLE_SHEET_URL:
        try:
            sql = f"SELECT * FROM product_list WHERE SKU='{sku}'"
            resp = requests.get(
                GOOGLE_SHEET_URL,
                params={"sql": sql},
                timeout=5,
            )
            resp.raise_for_status()
            data = resp.json()
            rows = data.get("rows", [])
            if rows:
                return _normalise(rows[0])
        except Exception:
            pass  # fall through to mock

    mock = _load_mock()
    raw = mock.get(sku)
    return _normalise(raw) if raw else None
