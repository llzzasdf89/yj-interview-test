"""
Order Model — reads order.xlsx and SKU.xlsx from project root.
"""
import os
from typing import Optional
from openpyxl import load_workbook

ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
ORDER_PATH = os.path.join(ROOT, "order.xlsx")
SKU_PATH = os.path.join(ROOT, "SKU.xlsx")


def _read_sheet(path: str, sheet_name: Optional[str] = None) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
        if any(v is not None for v in row)
    ]
    wb.close()
    return rows


def get_all_orders() -> list[dict]:
    return _read_sheet(ORDER_PATH)


def get_order_by_no(order_no: str) -> Optional[dict]:
    for order in get_all_orders():
        if order.get("OrderNo") == order_no:
            return order
    return None


def get_skus_by_order(order_no: str) -> list[dict]:
    all_skus = _read_sheet(SKU_PATH)
    return [s for s in all_skus if s.get("OrderNumber") == order_no]


def get_all_sku_rows() -> list[dict]:
    return _read_sheet(SKU_PATH)
